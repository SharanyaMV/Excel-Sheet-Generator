import pandas as pd
from io import BytesIO, StringIO
from typing import Dict, Any, List
from datetime import date, timedelta
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, Series
from openpyxl.styles import Font


FIRST_NAMES = [
    "Aarav", "Vivaan", "Aditya", "Arjun", "Sai", "Reyansh", "Krishna", "Ishaan",
    "Anaya", "Diya", "Aadhya", "Saanvi", "Priya", "Neha", "Kavya", "Meera",
    "Rahul", "Rohan", "Ananya", "Vikram", "Sneha", "Kiran", "Nisha", "Amit",
]
LAST_NAMES = [
    "Sharma", "Patel", "Reddy", "Nair", "Singh", "Kumar", "Gupta", "Joshi",
    "Mehta", "Rao", "Iyer", "Das", "Verma", "Kapoor", "Chopra", "Menon",
]
DEPARTMENTS = ["Engineering", "HR", "Finance", "Sales", "Marketing", "Operations", "IT", "Legal", "Support"]
JOB_TITLES = {
    "Engineering": ["Software Engineer", "Senior Developer", "QA Engineer", "DevOps Engineer"],
    "HR": ["HR Executive", "Recruiter", "HR Manager", "Training Coordinator"],
    "Finance": ["Accountant", "Financial Analyst", "Payroll Specialist", "Finance Manager"],
    "Sales": ["Sales Executive", "Account Manager", "Sales Lead", "Business Development Manager"],
    "Marketing": ["Marketing Executive", "Content Specialist", "SEO Analyst", "Brand Manager"],
    "Operations": ["Operations Executive", "Operations Manager", "Logistics Coordinator", "Process Analyst"],
    "IT": ["IT Support Engineer", "System Administrator", "Network Engineer", "IT Manager"],
    "Legal": ["Legal Associate", "Compliance Officer", "Contract Specialist", "Legal Manager"],
    "Support": ["Support Executive", "Customer Success Associate", "Support Lead", "Helpdesk Analyst"],
}
CITIES = [
    ("Bengaluru", "Karnataka", "India"),
    ("Hyderabad", "Telangana", "India"),
    ("Mumbai", "Maharashtra", "India"),
    ("Pune", "Maharashtra", "India"),
    ("Chennai", "Tamil Nadu", "India"),
    ("Delhi", "Delhi", "India"),
    ("Kolkata", "West Bengal", "India"),
    ("Ahmedabad", "Gujarat", "India"),
]
SUPPORTED_VISUALS = {"bar", "line", "pie", "scatter"}


def _full_name(index: int) -> str:
    return f"{FIRST_NAMES[index % len(FIRST_NAMES)]} {LAST_NAMES[(index * 3) % len(LAST_NAMES)]}"


def _first_name(index: int) -> str:
    return FIRST_NAMES[index % len(FIRST_NAMES)]


def _last_name(index: int) -> str:
    return LAST_NAMES[(index * 3) % len(LAST_NAMES)]


def _sample_values(name: str, col_type: str, rows: int):
    lower_name = name.lower()

    if lower_name in {"employee_id", "emp_id"} or lower_name.endswith("_id"):
        return [10000 + i + 1 for i in range(rows)]
    if lower_name in {"employee_name", "name", "full_name"}:
        return [_full_name(i) for i in range(rows)]
    if lower_name == "first_name":
        return [_first_name(i) for i in range(rows)]
    if lower_name == "last_name":
        return [_last_name(i) for i in range(rows)]
    if "manager_name" in lower_name:
        return [_full_name(i + 7) for i in range(rows)]
    if "emergency_contact_name" in lower_name:
        return [_full_name(i + 13) for i in range(rows)]
    if "gender" in lower_name:
        genders = ["Male", "Female", "Other"]
        return [genders[i % len(genders)] for i in range(rows)]
    if "date_of_birth" in lower_name or lower_name in {"dob", "birth_date"}:
        return [date(1970 + (i % 31), (i % 12) + 1, (i % 28) + 1) for i in range(rows)]
    if lower_name == "age":
        return [24 + (i % 36) for i in range(rows)]
    if "email" in lower_name:
        return [
            f"{_first_name(i).lower()}.{_last_name(i).lower()}{i + 1}@company.com"
            for i in range(rows)
        ]
    if "phone" in lower_name or "contact_number" in lower_name:
        return [f"+91-9{(800000000 + i * 137) % 1000000000:09d}" for i in range(rows)]
    if "address" in lower_name:
        return [f"{100 + i}, {['MG Road', 'Park Street', 'Brigade Road', 'Anna Salai'][i % 4]}" for i in range(rows)]
    if lower_name == "city":
        return [CITIES[i % len(CITIES)][0] for i in range(rows)]
    if lower_name == "state":
        return [CITIES[i % len(CITIES)][1] for i in range(rows)]
    if lower_name == "country":
        return [CITIES[i % len(CITIES)][2] for i in range(rows)]
    if "postal" in lower_name or "zip" in lower_name:
        return [f"{560000 + (i % 9000):06d}" for i in range(rows)]
    if "department" in lower_name:
        return [DEPARTMENTS[i % len(DEPARTMENTS)] for i in range(rows)]
    if "job_title" in lower_name or "designation" in lower_name or "title" in lower_name:
        values = []
        for i in range(rows):
            department = DEPARTMENTS[i % len(DEPARTMENTS)]
            titles = JOB_TITLES[department]
            values.append(titles[i % len(titles)])
        return values
    if "employment_type" in lower_name:
        types = ["Full-time", "Part-time", "Contract", "Intern"]
        return [types[i % len(types)] for i in range(rows)]
    if "joining_date" in lower_name or "hire_date" in lower_name:
        return [date(2015 + (i % 10), (i % 12) + 1, (i % 28) + 1) for i in range(rows)]
    if "salary" in lower_name:
        return [float(350000 + (i % 80) * 12500) for i in range(rows)]
    if "bonus" in lower_name:
        return [float(10000 + (i % 20) * 2500) for i in range(rows)]
    if "work_location" in lower_name:
        return [CITIES[(i + 2) % len(CITIES)][0] for i in range(rows)]
    if "shift" in lower_name:
        shifts = ["Morning", "General", "Evening", "Night"]
        return [shifts[i % len(shifts)] for i in range(rows)]
    if "employee_status" in lower_name or lower_name == "status":
        statuses = ["Active", "On Leave", "Probation", "Resigned"]
        return [statuses[i % len(statuses)] for i in range(rows)]
    if "experience" in lower_name:
        return [1 + (i % 25) for i in range(rows)]
    if "education" in lower_name:
        values = ["B.Tech", "MBA", "B.Com", "M.Tech", "B.Sc", "MCA"]
        return [values[i % len(values)] for i in range(rows)]
    if "skills" in lower_name:
        values = ["Python, SQL", "Excel, Reporting", "Java, Spring", "Recruitment", "Accounting", "Customer Support"]
        return [values[i % len(values)] for i in range(rows)]
    if "performance_rating" in lower_name or "rating" in lower_name:
        return [1 + (i % 5) for i in range(rows)]

    if col_type == "integer":
        return [i + 1 for i in range(rows)]
    if col_type == "float":
        return [round((i + 1) * 100.25, 2) for i in range(rows)]
    if col_type == "date":
        start = date.today()
        return [start + timedelta(days=i) for i in range(rows)]
    if col_type == "boolean":
        return [i % 2 == 0 for i in range(rows)]

    if "name" in lower_name:
        return [_full_name(i) for i in range(rows)]
    if "status" in lower_name:
        statuses = ["Pending", "In Progress", "Completed"]
        return [statuses[i % len(statuses)] for i in range(rows)]
    if lower_name.startswith("employee_field_"):
        return [f"Employee detail {i + 1}" for i in range(rows)]

    return [f"{name}_{i + 1}" for i in range(rows)]


def _sheet_title(title: str) -> str:
    cleaned = "".join(char for char in title if char not in r'[]:*?/\\')[:31]
    return cleaned or "DataSheet"


def _numeric_columns(df: pd.DataFrame) -> List[str]:
    return [
        column for column in df.columns
        if pd.api.types.is_numeric_dtype(df[column]) and not pd.api.types.is_bool_dtype(df[column])
    ]


def _category_columns(df: pd.DataFrame) -> List[str]:
    return [
        column for column in df.columns
        if not pd.api.types.is_numeric_dtype(df[column]) and not pd.api.types.is_datetime64_any_dtype(df[column])
    ]


def _default_visuals(df: pd.DataFrame) -> List[str]:
    numeric_columns = _numeric_columns(df)
    category_columns = _category_columns(df)
    visuals = []

    if category_columns:
        visuals.append("bar")
        visuals.append("pie")
    if numeric_columns:
        visuals.append("line")
    if len(numeric_columns) >= 2:
        visuals.append("scatter")

    return visuals or ["bar"]


def _write_table(ws, start_row: int, start_col: int, headers: List[str], rows: List[List[Any]]) -> int:
    for col_offset, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + col_offset, value=header)
        cell.font = Font(bold=True)

    for row_offset, row_values in enumerate(rows, start=1):
        for col_offset, value in enumerate(row_values):
            ws.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)

    return start_row + len(rows)


def _category_summary(df: pd.DataFrame, category_col: str, numeric_col: str = "") -> pd.DataFrame:
    if numeric_col:
        summary = (
            df.groupby(category_col, dropna=False)[numeric_col]
            .sum()
            .reset_index()
            .rename(columns={numeric_col: f"Total {numeric_col}"})
        )
    else:
        summary = (
            df[category_col]
            .value_counts(dropna=False)
            .rename_axis(category_col)
            .reset_index(name="Count")
        )

    return summary.head(12)


def _add_bar_chart(ws, df: pd.DataFrame, anchor: str, table_row: int) -> int:
    category_columns = _category_columns(df)
    numeric_columns = _numeric_columns(df)
    category_col = category_columns[0] if category_columns else df.columns[0]
    numeric_col = numeric_columns[0] if numeric_columns else ""
    summary = _category_summary(df, category_col, numeric_col)

    end_row = _write_table(ws, table_row, 1, list(summary.columns), summary.values.tolist())
    chart = BarChart()
    chart.title = f"{summary.columns[1]} by {category_col}"
    chart.y_axis.title = str(summary.columns[1])
    chart.x_axis.title = category_col
    chart.add_data(Reference(ws, min_col=2, min_row=table_row, max_row=end_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=table_row + 1, max_row=end_row))
    chart.height = 7
    chart.width = 12
    ws.add_chart(chart, anchor)
    return end_row + 3


def _add_pie_chart(ws, df: pd.DataFrame, anchor: str, table_row: int) -> int:
    category_columns = _category_columns(df)
    category_col = category_columns[0] if category_columns else df.columns[0]
    summary = _category_summary(df, category_col)

    end_row = _write_table(ws, table_row, 1, list(summary.columns), summary.values.tolist())
    chart = PieChart()
    chart.title = f"{category_col} Distribution"
    chart.add_data(Reference(ws, min_col=2, min_row=table_row, max_row=end_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=table_row + 1, max_row=end_row))
    chart.height = 7
    chart.width = 10
    ws.add_chart(chart, anchor)
    return end_row + 3


def _add_line_chart(ws, df: pd.DataFrame, anchor: str, table_row: int) -> int:
    numeric_columns = _numeric_columns(df)
    if not numeric_columns:
        return _add_bar_chart(ws, df, anchor, table_row)

    label_col = next(
        (column for column in df.columns if column not in numeric_columns),
        df.columns[0],
    )
    value_col = numeric_columns[0]
    rows = df[[label_col, value_col]].head(25).values.tolist()
    end_row = _write_table(ws, table_row, 1, [label_col, value_col], rows)

    chart = LineChart()
    chart.title = f"{value_col} Trend"
    chart.y_axis.title = value_col
    chart.x_axis.title = label_col
    chart.add_data(Reference(ws, min_col=2, min_row=table_row, max_row=end_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=table_row + 1, max_row=end_row))
    chart.height = 7
    chart.width = 12
    ws.add_chart(chart, anchor)
    return end_row + 3


def _add_scatter_chart(ws, df: pd.DataFrame, anchor: str, table_row: int) -> int:
    numeric_columns = _numeric_columns(df)
    if len(numeric_columns) < 2:
        return _add_line_chart(ws, df, anchor, table_row)

    x_col, y_col = numeric_columns[:2]
    rows = df[[x_col, y_col]].head(25).values.tolist()
    end_row = _write_table(ws, table_row, 1, [x_col, y_col], rows)

    chart = ScatterChart()
    chart.title = f"{y_col} vs {x_col}"
    chart.x_axis.title = x_col
    chart.y_axis.title = y_col
    x_values = Reference(ws, min_col=1, min_row=table_row + 1, max_row=end_row)
    y_values = Reference(ws, min_col=2, min_row=table_row + 1, max_row=end_row)
    chart.series.append(Series(y_values, x_values, title=y_col))
    chart.height = 7
    chart.width = 12
    ws.add_chart(chart, anchor)
    return end_row + 3


def _add_visuals_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, requested_visuals: List[str]) -> List[str]:
    workbook = writer.book
    ws = workbook.create_sheet("Visuals", 1)
    ws["A1"] = "Visual Summary"
    ws["A1"].style = "Title"
    ws["A2"] = "Charts are generated from the data sheet."

    visuals = [visual for visual in requested_visuals if visual in SUPPORTED_VISUALS]
    if not visuals:
        visuals = _default_visuals(df)

    chart_builders = {
        "bar": _add_bar_chart,
        "line": _add_line_chart,
        "pie": _add_pie_chart,
        "scatter": _add_scatter_chart,
    }
    anchors = ["D4", "D20", "D36", "D52"]
    table_row = 4
    generated = []

    for index, visual in enumerate(visuals[:4]):
        table_row = chart_builders[visual](ws, df, anchors[index], table_row)
        generated.append(visual)

    for column_letter in ("A", "B", "C"):
        ws.column_dimensions[column_letter].width = 22

    return generated


def generate_excel(schema_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Args:
        schema_data: Dictionary containing sheet_name, columns, and rows
    """
    try:
        sheet_name = schema_data.get("sheet_name", "DataSheet")
        columns = schema_data.get("columns", [])
        rows = schema_data.get("rows", 10)
        output_formats = schema_data.get("output_formats", ["excel"])
        requested_visuals = schema_data.get("visuals", [])

        # 1. Generate sample data
        data = {}
        for column in columns:
            name = column["name"]
            col_type = column.get("type", "string")
            data[name] = _sample_values(name, col_type, rows)

        df = pd.DataFrame(data)

        # 2. Generate the requested file formats in memory.
        artifacts = []
        generated_visuals = []
        if "excel" in output_formats:
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                safe_sheet_name = _sheet_title(sheet_name)
                df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                generated_visuals = _add_visuals_sheet(writer, df, requested_visuals)
            artifacts.append(
                {
                    "file_name": f"{sheet_name}.xlsx",
                    "content": excel_buffer.getvalue(),
                    "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                }
            )

        if "csv" in output_formats:
            csv_buffer = StringIO()
            df.to_csv(csv_buffer, index=False)
            artifacts.append(
                {
                    "file_name": f"{sheet_name}.csv",
                    "content": csv_buffer.getvalue().encode("utf-8-sig"),
                    "content_type": "text/csv",
                }
            )

        if not artifacts:
            raise ValueError("No supported output format requested. Use excel, csv, or both.")

        file_names = [artifact["file_name"] for artifact in artifacts]

        return {
            "status": "success",
            "file_name": file_names[0],
            "file_names": file_names,
            "message": f"Generated {', '.join(file_names)} successfully.",
            "artifacts": artifacts,
            "columns": [col["name"] for col in columns],
            "row_count": rows,
            "visuals": generated_visuals,
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}
